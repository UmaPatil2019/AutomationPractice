import openpyxl
file = "/Users/maheshpatil/Library/Mobile Documents/com~apple~Numbers/Documents/File1.numbers"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

#wite data
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="welcome"

#save file
workbook.save(file)

#how to add different data into excel file
sheet.cell(1,1) = 123
sheet.cell(1,2)= "uma"
sheet.cell(1,3)= "patil"

sheet.cell(2,1) = 234
sheet.cell(2,2)= "a"
sheet.cell(2,3)= "b"

sheet.cell(3,1) = 345
sheet.cell(3,2)= "c"
sheet.cell(3,3)= "d"

#save filel
workbook.save(file)