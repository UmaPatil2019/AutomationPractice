from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.select import Select
import time

#excel file -> WorkBook -> Sheets-> Rows-> Columns/cells
workbook = load_workbook("/Users/maheshpatil/Downloads/Financial Sample.xlsx")
sheet = workbook["Sheet1"]

#count number of rows
rows = sheet.max_row
#count number of columns
cols = sheet.max_column

#read all row and col
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value, end='      ')
    print()



